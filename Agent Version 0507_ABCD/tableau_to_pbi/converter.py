"""High-level orchestrator.

Pipeline (always linear, no branching beyond the TWB/TWBX flag):

    parse  -> read .twb XML into a clean IR dict
    [opt]  -> dump IR JSON for inspection
    model  -> build TMDL semantic model
              (full data model for .twbx, stub-only for .twb)
    report -> build pages and visuals from the IR
    write  -> drop everything onto disk in PBIP layout

Behavior split between input types — per project direction:

    .twb   input:  *only* visual creation. Data model is a stub
                   placeholder so the PBIP loads; visual queries come
                   out empty but layout is preserved.

    .twbx  input:  full data model (tables, columns, relationships from
                   the workbook's <object-graph>) plus visuals with
                   queries wired to those tables/columns.
"""

import os
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .hyper  import HyperData, HyperRegistry
from .model  import SemanticModel
from .parser import TWBParser
from .report import ReportBuilder
from .utils  import write_json
from .writer import PBIPWriter


# ---------------------------------------------------------------------------
# Converter
# ---------------------------------------------------------------------------

class Converter:
    def __init__(
        self,
        twb_path:    str,
        output:      Optional[str] = None,
        stub_only:   bool = False,
        debug_ir:    bool = True,
        hyper_paths: Optional[List[str]] = None,
        hints:       Optional[Dict[str, Dict[str, List[str]]]] = None,
    ):
        self.twb_path    = twb_path
        self.stub_only   = stub_only
        self.debug_ir    = debug_ir
        self.hyper_paths = hyper_paths or []
        # Field-resolution hints supplied by the agent codebase. Empty
        # dict (default) means the converter behaves identically to a
        # no-agent run.
        self.hints       = hints or {}

        if output:
            self.output = Path(output)
        else:
            stem        = Path(twb_path).stem
            self.output = Path(twb_path).resolve().parent / f"{stem}_pbip"

    def _bind_hyper_files(
        self, datasources: List[Dict[str, Any]],
    ) -> Dict[str, HyperData]:
        """Pair each parsed datasource with the .hyper file it owns.

        Delegates to HyperRegistry, which honours the <connection
        class='hyper' dbname='...'> bindings declared inside each
        datasource. Returning {} when no hyper files were supplied or
        none could be loaded is fine — the model will fall back to the
        coarser TWB-XML column types.

        Side effect: `self._hyper_registry` is kept around so the
        mapping report can read which path got bound to which ds.
        """
        if not self.hyper_paths:
            self._hyper_registry: Optional[HyperRegistry] = None
            return {}
        self._hyper_registry = HyperRegistry(self.hyper_paths)
        return self._hyper_registry.bind(datasources)

    def _write_mapping_report(
        self,
        datasources: List[Dict[str, Any]],
        model:       "SemanticModel",
    ) -> None:
        """Write datasource_mapping.xlsx to the output dir.

        Three sheets capture the full TWB -> PBIP wiring:

          Datasources    one row per Tableau datasource: caption, hyper file,
                         declared extracts, count of TMDL tables produced.
          Columns        one row per PBI column: datasource, TMDL table,
                         PBI column name, Tableau column reference (the
                         original [TableCaption!Col]-style ref the .twb XML
                         used), and the hyper/CSV header that the column
                         actually binds to. This is the file a user opens
                         when a visual seems to reference the wrong column
                         or when a calculation references a header that
                         doesn't exist after the hyper extract is loaded.
          Relationships  one row per emitted relationship — these come
                         exclusively from <object-graph><relationship>
                         blocks in the .twb. The converter does NOT infer
                         relationships from common column names, so this
                         sheet should match what a user would see in the
                         Tableau "Data" tab "noodle" diagram.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            print("  [MAP] openpyxl not available — skipping Excel mapping.")
            return

        bound_paths: Dict[str, str] = {}
        if getattr(self, "_hyper_registry", None) is not None:
            bound_paths = dict(self._hyper_registry.bound_paths_by_ds)

        wb = Workbook()
        # Default sheet -> Datasources
        ws_ds = wb.active
        ws_ds.title = "Datasources"
        ws_ds.append([
            "Datasource", "Caption", "Hyper File",
            "Declared Extracts", "TMDL Tables",
        ])
        for ds in datasources:
            ds_name = ds["name"]
            hyper_file = (Path(bound_paths[ds_name]).name
                          if ds_name in bound_paths else "")
            extracts = "; ".join(ds.get("extracts", []) or [])
            tmdl_tables = [t["name"] for t in model.tables
                           if t.get("datasource") == ds_name]
            ws_ds.append([
                ds_name, ds.get("caption", ""), hyper_file,
                extracts, "; ".join(tmdl_tables),
            ])

        # Columns sheet — the load-bearing one for diagnosing mismatches
        # between TWB column refs and the hyper/CSV headers PBI binds to.
        ws_col = wb.create_sheet("Columns")
        ws_col.append([
            "Datasource",
            "TMDL Table",
            "PBI Column Name",
            "Tableau Column Reference",
            "Hyper / CSV Header",
            "Data Type",
            "Hidden",
        ])
        for t in model.tables:
            ds_name  = t.get("datasource", "")
            tbl_name = t["name"]
            for col in t.get("columns", []):
                ws_col.append([
                    ds_name,
                    tbl_name,
                    col["name"],
                    col.get("tableauRef", col.get("sourceCol", "")),
                    col.get("sourceCol", ""),
                    col.get("tmdlType", ""),
                    "Yes" if col.get("hidden") else "",
                ])

        # Relationships sheet — only TWB-declared rels, never inferred.
        ws_rel = wb.create_sheet("Relationships")
        ws_rel.append([
            "From Table", "From Column", "To Table", "To Column",
            "Active", "Source",
        ])
        for r in model.relationships:
            ws_rel.append([
                r["fromTable"], r["fromColumn"],
                r["toTable"],   r["toColumn"],
                "Yes" if r.get("isActive", True) else "No",
                "TWB <object-graph><relationship>",
            ])

        # Skipped Relationships sheet — TWB-declared rels that the
        # converter could not emit (missing endpoint, self-join, etc.).
        # Surfaced so the user can confirm nothing important is missing.
        ws_skip = wb.create_sheet("Skipped Relationships")
        ws_skip.append([
            "Datasource", "From Table", "From Column",
            "To Table", "To Column", "Reason",
        ])
        for r in getattr(model, "skipped_relationships", []) or []:
            ws_skip.append([
                r.get("datasource", ""),
                r.get("fromTable", ""), r.get("fromColumn", ""),
                r.get("toTable", ""),   r.get("toColumn", ""),
                r.get("reason", ""),
            ])

        # Auto-width pass (rough): set width to max content length per col.
        for sheet in (ws_ds, ws_col, ws_rel, ws_skip):
            for col_cells in sheet.columns:
                max_len = 0
                letter = col_cells[0].column_letter
                for cell in col_cells:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                sheet.column_dimensions[letter].width = min(max_len + 2, 60)

        out_path = self.output / "datasource_mapping.xlsx"
        wb.save(out_path)
        print(f"  [MAP] datasource_mapping.xlsx -> {out_path}")

    def run(self) -> None:
        name = Path(self.twb_path).stem
        mode = "TWB (visuals only, stub model)" if self.stub_only else "TWBX (full model + visuals)"
        print(f"Tableau -> PBIP  |  {name}")
        print(f"  Mode:    {mode}")
        print(f"  Source:  {self.twb_path}")
        print(f"  Output:  {self.output}")

        # ---- 1. parse ----------------------------------------------------
        print("  [1/4] Parsing Tableau workbook...")
        parser = TWBParser(self.twb_path)
        parser.parse()
        print(f"        datasources={len(parser.datasources)}  "
              f"parameters={len(parser.parameters)}  "
              f"worksheets={len(parser.worksheets)}  "
              f"dashboards={len(parser.dashboards)}")

        # Optionally dump IR JSON. This is a debug aid — useful when a
        # visual ends up looking wrong and the user wants to see exactly
        # what the parser pulled out.
        if self.debug_ir:
            self.output.mkdir(parents=True, exist_ok=True)
            write_json(self.output / "_ir.json", _build_ir(parser))

# ---- 2. model ----------------------------------------------------
        print("  [2/4] Building semantic model...")
        hyper_data_by_ds = self._bind_hyper_files(parser.datasources)
        model = SemanticModel(
            parser.datasources,
            parameters=parser.parameters,
            stub_only=self.stub_only,
            hyper_data_by_ds=hyper_data_by_ds,
            worksheets=parser.worksheets,
        )
        model.build()
        total_cols = sum(len(t["columns"]) for t in model.tables)
        total_measures = sum(len(t.get("measures", [])) for t in model.tables)
        param_table = any(t["name"] == "Parameters" for t in model.tables)
        print(f"        tables={len(model.tables)}  columns={total_cols}  "
              f"relationships={len(model.relationships)}  "
              f"measures={total_measures}  params={'yes' if param_table else 'no'}")

        # ---- 3. report ---------------------------------------------------
        print("  [3/4] Building report pages and visuals...")
        rb = ReportBuilder(parser.datasources, parser.worksheets,
                          parser.dashboards, model, hints=self.hints)
        pages = rb.build()
        from collections import Counter
        type_dist = Counter(
            v["visual"]["visualType"] for p in pages for v in p.get("visuals", [])
        )
        print(f"        pages={len(pages)}  "
              f"visuals={sum(len(p.get('visuals', [])) for p in pages)}")
        print(f"        types={dict(type_dist)}")

        # ---- 4. write ----------------------------------------------------
        print("  [4/4] Writing PBIP files...")
        PBIPWriter(self.output, name).write(model, pages)

        # Write datasource mapping report (helps debug visual field issues)
        self._write_mapping_report(parser.datasources, model)

        print(f"  Done. Open {self.output / (name + '.pbip')} in Power BI Desktop.")


# ---------------------------------------------------------------------------
# Intermediate representation (parser dicts -> JSON-friendly snapshot)
# ---------------------------------------------------------------------------

def _build_ir(parser: TWBParser) -> Dict[str, Any]:
    """Build a JSON-serializable view of the parsed workbook.

    Mostly a 1:1 copy of the parser's dicts — written to disk so the user
    can inspect what was extracted and see why a visual landed where it
    did. This is the "XML -> JSON" stage the user asked about; everything
    downstream of here is just JSON-to-PBIP transformation.
    """
    return {
        "source_file": parser.twb_path,
        "datasources": [
            {
                "name":          ds["name"],
                "caption":       ds["caption"],
                "connection":    ds["connection"],
                "extracts":      ds.get("extracts", []),
                # Stringify (table, col) tuples so json.dump doesn't fail.
                "colsMap":       {k: list(v) for k, v in
                                  (ds.get("colsMap") or {}).items()},
                "objects":       ds["objects"],
                "columns":       ds["columns"],
                "relationships": ds["relationships"],
            }
            for ds in parser.datasources
        ],
        "parameters": parser.parameters,
        "worksheets": parser.worksheets,
        "dashboards": parser.dashboards,
    }


# ---------------------------------------------------------------------------
# .twbx support
# ---------------------------------------------------------------------------

def extract_twbx(
    twbx_path: str, dest_dir: str,
) -> Tuple[str, List[str]]:
    """Pull the .twb and all .hyper files from a .twbx archive.

    Returns (twb_path, [hyper_path, ...]).
    """
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)

    twb_path:    Optional[str] = None
    hyper_paths: List[str]     = []

    with zipfile.ZipFile(twbx_path, "r") as zf:
        for member in zf.namelist():
            if member.endswith(".twb"):
                zf.extract(member, dest_dir)
                twb_path = str(dest / member)
            elif member.endswith(".hyper"):
                zf.extract(member, dest_dir)
                hyper_paths.append(str(dest / member))

    if not twb_path:
        raise ValueError(f"No .twb found inside {twbx_path}")

    if hyper_paths:
        print(f"  [HYPER] Found {len(hyper_paths)} hyper file(s): "
              + ", ".join(Path(p).name for p in hyper_paths))

    return twb_path, hyper_paths


# ---------------------------------------------------------------------------
# CLI / public entry point
# ---------------------------------------------------------------------------

def run(
    input_path: Optional[str] = None,
    output:     Optional[str] = None,
    *,
    twb:  Optional[str] = None,   # legacy
    twbx: Optional[str] = None,   # legacy
) -> None:
    """Convert a Tableau workbook to PBIP.

    The mode is auto-detected from the file extension:
        .twb   -> visuals-only (stub data model)
        .twbx  -> full data model + visuals

    Either pass `input_path=` directly, or use the legacy `twb=` /
    `twbx=` keyword arguments (kept for the existing CLI shim).
    """
    src = input_path or twbx or twb
    if not src:
        raise ValueError("No input file given.")
    if not os.path.exists(src):
        raise FileNotFoundError(f"Input not found: {src}")

    ext = Path(src).suffix.lower()
    if ext == ".twbx":
        stem             = Path(src).stem
        work_dir         = str(Path(src).parent / f"_{stem}_extracted")
        twb_path, hypers = extract_twbx(src, work_dir)
        out_path         = output or str(Path(src).parent / f"{stem}_pbip")
        Converter(twb_path, out_path, stub_only=False, hyper_paths=hypers).run()
    elif ext == ".twb":
        Converter(src, output, stub_only=True).run()
    else:
        raise ValueError(f"Unrecognized extension {ext!r} (expected .twb or .twbx).")
