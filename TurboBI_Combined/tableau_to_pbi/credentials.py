"""Credentials loader for Tableau -> PBIP conversion.

Reads a JSON or XLSX credentials file and matches entries to the
connection metadata already parsed from the .twb XML.  Two things
happen with a matched credential:

  1. Connection overrides — server / database / port / schema fields
     in the matched entry replace the values the parser read from the
     .twb.  This lets you promote a workbook from a dev connection to
     a prod connection without touching the Tableau file.

  2. Credentials manifest — `converter.py` writes a
     ``credentials_manifest.json`` alongside the PBIP output.  That
     file documents the username (and whether a password was supplied)
     for each live data source so the user knows exactly what to enter
     in Power BI Desktop > Transform Data > Data source settings, or
     what to push via the Power BI REST API.

Passwords are intentionally NOT embedded in the emitted M expressions.
Power BI's TMDL / PBIP format has no secure in-file credential store;
embedding plaintext passwords in partition M would expose them in any
version-controlled repo.  The credentials manifest is the correct
hand-off point for an automated deployment script.

--------------------------------------------------------------------
JSON format
--------------------------------------------------------------------
    {
      "connections": [
        {
          "class":    "sqlserver",
          "server":   "prod-sql.mycompany.com",
          "database": "AnalyticsDB",
          "port":     "1433",
          "schema":   "dbo",
          "username": "svc_pbi",
          "password": "s3cr3t"
        },
        {
          "class":    "postgres",
          "server":   "prod-pg.mycompany.com",
          "database": "analytics",
          "port":     "5432",
          "schema":   "public",
          "username": "pbi_reader",
          "password": "s3cr3t"
        },
        {
          "class":     "snowflake",
          "server":    "myaccount.snowflakecomputing.com",
          "warehouse": "COMPUTE_WH",
          "database":  "ANALYTICS",
          "schema":    "PUBLIC",
          "username":  "PBI_USER",
          "password":  "s3cr3t"
        }
      ]
    }

  ``class`` must match the Tableau connection class value exactly
  (case-insensitive).  ``server`` is optional — when omitted the entry
  matches any server of that class (useful for single-server workbooks).
  When both ``class`` + ``server`` match a more specific entry is
  preferred over a class-only entry.

--------------------------------------------------------------------
XLSX format
--------------------------------------------------------------------
  Create a sheet named **Credentials** (or use the first sheet).
  The first row must be column headers (case-insensitive).  Recognised
  headers: class, server, database, port, schema, warehouse,
  username, password.  Any extra headers are stored as additional
  override keys and merged into the connection dict.

  Example:

    | class     | server                    | database    | username   | password |
    |-----------|---------------------------|-------------|------------|----------|
    | sqlserver | prod-sql.mycompany.com    | AnalyticsDB | svc_pbi    | s3cr3t   |
    | postgres  | prod-pg.mycompany.com     | analytics   | pbi_reader | s3cr3t   |
    | snowflake | myacct.snowflake.com      | ANALYTICS   | PBI_USER   | s3cr3t   |
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from ._logging import get_logger

_log_creds = get_logger("CREDS")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

class CredentialEntry:
    """One row of the credentials file."""

    _IDENTITY_KEYS = frozenset({
        "class", "server", "host", "database", "dbname", "db",
        "port", "schema", "warehouse",

        # Datasource-level matching
        "datasource", "datasource_name", "datasource_caption", "caption",

        # Databricks fields
        "http_path", "httppath", "http path",
        "catalog",
        "connector_function", "connector",

        # Extract behavior
        "prefer_live_over_extract",

        # Custom SQL override — when present, the partition M emits
        # Value.NativeQuery(Source, "<query>") instead of the table-
        # navigation path. Lets a credentials file repoint a workbook
        # at a different query without editing the .twb XML.
        "query", "custom_sql", "sql",
    })

    _AUTH_KEYS = frozenset({
        "username", "user",
        "password", "pwd",
        "token", "pat", "access_token", "personal_access_token",
    })

    def __init__(self, data: Dict[str, Any]) -> None:
        def _s(key: str, *aliases: str) -> str:
            for k in (key, *aliases):
                v = data.get(k)
                if v is not None and str(v).strip():
                    return str(v).strip()
            return ""

        self.cls       = _s("class").lower()
        self.server    = _s("server", "host")
        self.database  = _s("database", "dbname", "db")
        self.port      = _s("port")
        self.schema    = _s("schema")
        self.warehouse = _s("warehouse")
        self.username  = _s("username", "user")
        self.password  = _s("password", "pwd")

        # Datasource-aware matching
        self.datasource = _s("datasource", "datasource_name")
        self.caption    = _s("datasource_caption", "caption")

        # Databricks-specific fields
        self.http_path = _s("http_path", "httppath", "http path")
        self.catalog   = _s("catalog")
        self.token     = _s("token", "pat", "access_token", "personal_access_token")
        self.connector_function = _s("connector_function", "connector")

        # Custom SQL override. Accepted under any of these header names so
        # operators can pick whichever reads best in their JSON/XLSX:
        #   "query"       — most common
        #   "custom_sql"  — matches the Tableau XML attribute spelling
        #   "sql"         — terse
        self.query     = _s("query", "custom_sql", "sql")

        prefer_live = _s("prefer_live_over_extract")
        self.prefer_live_over_extract = (
            prefer_live.strip().lower() in {"1", "true", "yes", "y"}
            if prefer_live
            else True
        )

        self.extra: Dict[str, str] = {
            k: str(v).strip()
            for k, v in data.items()
            if k.lower() not in (self._IDENTITY_KEYS | self._AUTH_KEYS)
            and v is not None
            and str(v).strip()
        }

    def apply_to(self, conn: Dict[str, Any]) -> Dict[str, Any]:
        """Return a copy of conn with connection overrides applied.

        Security note:
        Passwords and PATs are intentionally not copied into the M connection.
        """
        result = dict(conn)

        # Important: allow federated/extract Tableau connection to become
        # Databricks when credentials explicitly say class=databricks.
        if self.cls:
            result["class"] = self.cls

        if self.server:
            result["server"] = self.server
        if self.database:
            result["dbname"] = self.database
        if self.port:
            result["port"] = self.port
        if self.schema:
            result["schema"] = self.schema
        if self.warehouse:
            result["warehouse"] = self.warehouse

        if self.http_path:
            result["http_path"] = self.http_path
        if self.catalog:
            result["catalog"] = self.catalog
        if self.connector_function:
            result["connector_function"] = self.connector_function

        # Custom SQL override flows through `conn` so `_render_partition_m`
        # can route it to Value.NativeQuery. Stored under the same key the
        # model already reads for Tableau-embedded custom SQL (`customSql`).
        # When the credentials entry supplies a query, it WINS over any
        # Tableau-embedded SQL — the credentials file is the override layer.
        if self.query:
            result["customSql"] = [{"name": "credentials override",
                                    "sql": self.query}]
            result["force_live_for_custom_sql"] = True

        result["prefer_live_over_extract"] = self.prefer_live_over_extract

        for k, v in self.extra.items():
            result[k] = v

        return result
    
class CredentialStore:
    """Ordered list of CredentialEntry objects."""

    def __init__(self, entries: List[CredentialEntry]) -> None:
        self._entries = entries

    def match(
        self,
        conn: Dict[str, Any],
        datasource: str = "",
        caption: str = "",
    ) -> Optional[CredentialEntry]:
        """Return the best-matching entry for a connection dict.

        Priority:
          1. datasource / caption match
          2. class + server exact match
          3. class-only match
          4. safe fallback for Tableau extract/federated connection when
             exactly one live Databricks credential exists
        """
        cls    = (conn.get("class") or "").strip().lower()
        server = (conn.get("server") or "").strip().lower()
        ds_lc  = (datasource or "").strip().lower()
        cap_lc = (caption or "").strip().lower()

        # 1. Datasource/caption match
        for e in self._entries:
            if e.datasource and e.datasource.lower() == ds_lc:
                return e
            if e.caption and e.caption.lower() == cap_lc:
                return e

        # 2. Class + server
        for e in self._entries:
            if e.cls == cls and e.server.lower() == server:
                return e

        # 3. Class only
        for e in self._entries:
            if e.cls == cls and not e.server:
                return e

        # 4. Safe fallback:
        # Tableau extract workbooks can expose class=federated/hyper/extract
        # even when user wants to redirect to Databricks.
        extract_like = cls in {"", "federated", "hyper", "extract"}
        live_databricks_entries = [
            e for e in self._entries
            if e.cls in {
                "databricks",
                "azure-databricks",
                "azuredatabricks",
                "databricks-sql",
                "spark-sql",
                "spark",
            }
            and e.server
            and e.http_path
        ]

        if extract_like and len(live_databricks_entries) == 1:
            return live_databricks_entries[0]

        return None

    def apply_overrides(
        self,
        conn: Dict[str, Any],
        datasource: str = "",
        caption: str = "",
    ) -> Dict[str, Any]:
        """Return conn with the matched entry's overrides applied."""
        entry = self.match(conn, datasource=datasource, caption=caption)
        return entry.apply_to(conn) if entry is not None else conn
    
    @property
    def entries(self) -> List[CredentialEntry]:
        return list(self._entries)


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_credentials(path: str) -> CredentialStore:
    """Load a credentials file and return a :class:`CredentialStore`.

    Supported formats: ``.json`` and ``.xlsx`` / ``.xls``.

    Raises :class:`FileNotFoundError` when the file is absent.
    Raises :class:`ValueError` for unrecognised extensions.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Credentials file not found: {path}")

    suffix = p.suffix.lower()
    if suffix == ".json":
        return _load_json(p)
    if suffix in (".xlsx", ".xls"):
        return _load_xlsx(p)
    raise ValueError(
        f"Unsupported credentials file format: {suffix!r}  (use .json or .xlsx)"
    )


def _load_json(path: Path) -> CredentialStore:
    with open(path, encoding="utf-8") as fh:
        raw = json.load(fh)

    # Accept both {"connections": [...]} and a bare list.
    if isinstance(raw, list):
        rows = raw
    else:
        rows = (
            raw.get("connections")
            or raw.get("credentials")
            or raw.get("datasources")
            or []
        )

    entries = [CredentialEntry(r) for r in rows if isinstance(r, dict)]
    _log_creds.info(f"Loaded {len(entries)} credential entries from {path.name}")
    return CredentialStore(entries)


def _load_xlsx(path: Path) -> CredentialStore:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read XLSX credentials. "
            "Install with:  pip install openpyxl"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)

    # Prefer a sheet named "Credentials"; fall back to the active sheet.
    ws = wb["Credentials"] if "Credentials" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return CredentialStore([])

    # First row = headers (normalise to lower-case stripped strings).
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    entries: List[CredentialEntry] = []
    for row in rows[1:]:
        # Skip fully-empty rows.
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        entry_data: Dict[str, Any] = {}
        for h, v in zip(headers, row):
            if h and v is not None and str(v).strip():
                entry_data[h] = str(v).strip()
        if entry_data:
            entries.append(CredentialEntry(entry_data))

    _log_creds.info(f"Loaded {len(entries)} credential entries from {path.name}")
    return CredentialStore(entries)