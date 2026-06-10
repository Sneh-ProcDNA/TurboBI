"""High-level orchestrator.

Pipeline (always linear, no branching beyond the TWB/TWBX flag):

    parse  -> read .twb XML into a clean IR dict
    [opt]  -> dump IR JSON for inspection
    model  -> build TMDL semantic model
              (full data model for .twbx, stub-only for .twb)
    report -> build pages and visuals from the IR
    write  -> drop everything onto disk in PBIP layout

Behavior split between input types — per project direction:

    .twb   input:  *only* visual creation. Data model is a stub
                   placeholder so the PBIP loads; visual queries come
                   out empty but layout is preserved.

    .twbx  input:  full data model (tables, columns, relationships from
                   the workbook's <object-graph>) plus visuals with
                   queries wired to those tables/columns.
"""

import json
import os
import zipfile
import shutil
import gc
import stat
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ._logging   import get_logger
from .hyper       import HyperData, HyperRegistry
from .model       import SemanticModel
from .parser      import TWBParser
from .report      import ReportBuilder
from .utils       import write_json
from .writer      import PBIPWriter

_log_map   = get_logger("MAP")
_log_clean = get_logger("CLEAN")
_log_creds = get_logger("CREDS")
_log_hyper = get_logger("HYPER")


# ---------------------------------------------------------------------------
# Converter
# ---------------------------------------------------------------------------

class Converter:
    def __init__(
        self,
        twb_path: str,
        output: Optional[str] = None,
        stub_only: bool = False,
        debug_ir: bool = True,
        hyper_paths: Optional[List[str]] = None,
        hints: Optional[Dict[str, Dict[str, List[str]]]] = None,
        credentials_path: Optional[str] = None,
        clean_output: bool = True,
    ):
        self.twb_path    = twb_path
        self.stub_only   = stub_only
        self.debug_ir    = debug_ir
        self.hyper_paths = hyper_paths or []
        # Field-resolution hints supplied by the agent codebase. Empty
        # dict (default) means the converter behaves identically to a
        # no-agent run.
        self.hints       = hints or {}

        # Optional credentials file (JSON or XLSX).  When provided the
        # credential store is used to:
        #   (a) override connection fields in partition-M expressions
        #       (server / database / port / schema promotion), and
        #   (b) write a credentials_manifest.json alongside the output
        #       so the user knows what to configure in PBI Desktop or
        #       via the Power BI REST API.
        self._credential_store = None
        if credentials_path:
            from .credentials import load_credentials
            self._credential_store = load_credentials(credentials_path)

        if output:
            self.output = Path(output)
        else:
            stem        = Path(twb_path).stem
            self.output = Path(twb_path).resolve().parent / f"{stem}_pbip"

    def _bind_hyper_files(
        self, datasources: List[Dict[str, Any]],
    ) -> Dict[str, HyperData]:
        """Pair each parsed datasource with the .hyper file it owns.

        Delegates to HyperRegistry, which honours the <connection
        class='hyper' dbname='...'> bindings declared inside each
        datasource. Returning {} when no hyper files were supplied or
        none could be loaded is fine — the model will fall back to the
        coarser TWB-XML column types.

        Side effect: `self._hyper_registry` is kept around so the
        mapping report can read which path got bound to which ds.
        """
        if not self.hyper_paths:
            self._hyper_registry: Optional[HyperRegistry] = None
            return {}
        self._hyper_registry = HyperRegistry(self.hyper_paths)
        return self._hyper_registry.bind(datasources)

    def _write_mapping_report(
        self,
        datasources: List[Dict[str, Any]],
        model:       "SemanticModel",
    ) -> None:
        """Write datasource_mapping.xlsx to the output dir.

        Three sheets capture the full TWB -> PBIP wiring:

          Datasources    one row per Tableau datasource: caption, hyper file,
                         declared extracts, count of TMDL tables produced.
          Columns        one row per PBI column: datasource, TMDL table,
                         PBI column name, Tableau column reference (the
                         original [TableCaption!Col]-style ref the .twb XML
                         used), and the hyper/CSV header that the column
                         actually binds to. This is the file a user opens
                         when a visual seems to reference the wrong column
                         or when a calculation references a header that
                         doesn't exist after the hyper extract is loaded.
          Relationships  one row per emitted relationship — these come
                         exclusively from <object-graph><relationship>
                         blocks in the .twb. The converter does NOT infer
                         relationships from common column names, so this
                         sheet should match what a user would see in the
                         Tableau "Data" tab "noodle" diagram.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            _log_map.warning("openpyxl not available — skipping Excel mapping.")
            return

        bound_paths: Dict[str, str] = {}
        if getattr(self, "_hyper_registry", None) is not None:
            bound_paths = dict(self._hyper_registry.bound_paths_by_ds)

        wb = Workbook()
        # Default sheet -> Datasources
        ws_ds = wb.active
        ws_ds.title = "Datasources"
        ws_ds.append([
            "Datasource", "Caption", "Hyper File",
            "Declared Extracts", "TMDL Tables",
        ])
        for ds in datasources:
            ds_name = ds["name"]
            hyper_file = (Path(bound_paths[ds_name]).name
                          if ds_name in bound_paths else "")
            extracts = "; ".join(ds.get("extracts", []) or [])
            tmdl_tables = [t["name"] for t in model.tables
                           if t.get("datasource") == ds_name]
            ws_ds.append([
                ds_name, ds.get("caption", ""), hyper_file,
                extracts, "; ".join(tmdl_tables),
            ])

        # Columns sheet — the load-bearing one for diagnosing mismatches
        # between TWB column refs and the hyper/CSV headers PBI binds to.
        ws_col = wb.create_sheet("Columns")
        ws_col.append([
            "Datasource",
            "TMDL Table",
            "PBI Column Name",
            "Tableau Column Reference",
            "Hyper / CSV Header",
            "Data Type",
            "Hidden",
        ])
        for t in model.tables:
            ds_name  = t.get("datasource", "")
            tbl_name = t["name"]
            for col in t.get("columns", []):
                ws_col.append([
                    ds_name,
                    tbl_name,
                    col["name"],
                    col.get("tableauRef", col.get("sourceCol", "")),
                    col.get("sourceCol", ""),
                    col.get("tmdlType", ""),
                    "Yes" if col.get("hidden") else "",
                ])

        # Relationships sheet — only TWB-declared rels, never inferred.
        ws_rel = wb.create_sheet("Relationships")
        ws_rel.append([
            "From Table", "From Column", "To Table", "To Column",
            "Active", "Source",
        ])
        for r in model.relationships:
            ws_rel.append([
                r["fromTable"], r["fromColumn"],
                r["toTable"],   r["toColumn"],
                "Yes" if r.get("isActive", True) else "No",
                "TWB <object-graph><relationship>",
            ])

        # Skipped Relationships sheet — TWB-declared rels that the
        # converter could not emit (missing endpoint, self-join, etc.).
        # Surfaced so the user can confirm nothing important is missing.
        ws_skip = wb.create_sheet("Skipped Relationships")
        ws_skip.append([
            "Datasource", "From Table", "From Column",
            "To Table", "To Column", "Reason",
        ])
        for r in getattr(model, "skipped_relationships", []) or []:
            ws_skip.append([
                r.get("datasource", ""),
                r.get("fromTable", ""), r.get("fromColumn", ""),
                r.get("toTable", ""),   r.get("toColumn", ""),
                r.get("reason", ""),
            ])

        # Auto-width pass (rough): set width to max content length per col.
        for sheet in (ws_ds, ws_col, ws_rel, ws_skip):
            for col_cells in sheet.columns:
                max_len = 0
                letter = col_cells[0].column_letter
                for cell in col_cells:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                sheet.column_dimensions[letter].width = min(max_len + 2, 60)

        out_path = self.output / "datasource_mapping.xlsx"
        wb.save(out_path)
        _log_map.info(f"datasource_mapping.xlsx -> {out_path}")

    def _write_credentials_manifest(
        self,
        datasources: List[Dict[str, Any]],
    ) -> None:
        """Write credentials_manifest.json alongside the PBIP output.

        The manifest lists every live data source together with the
        connection parameters and (if a credentials file was supplied)
        the username that should be configured in Power BI.

        Passwords are deliberately omitted from this file — only a
        boolean ``has_password`` flag indicates that one was present in
        the credentials file.  The manifest is intended as an input for
        a deployment script that calls the Power BI REST API to set
        credentials, not as a plaintext password store.

        Power BI Desktop users: open the PBIP, then go to
        Transform Data > Data source settings and enter the credentials
        listed here for each connection string.
        """
        manifest: Dict[str, Any] = {
            "note": (
                "Set these credentials in Power BI Desktop via "
                "Transform Data > Data source settings, or use the "
                "Power BI REST API datasets/{id}/datasources endpoint."
            ),
            "datasources": [],
        }

        _SKIP_CLASSES = {"federated", "hyper", "extract", ""}

        for ds in datasources:
            conn = ds.get("connection") or {}
            cls  = (conn.get("class") or "").strip().lower()
            if cls in _SKIP_CLASSES:
                continue

            # Apply credential overrides so the manifest reflects the
            # *effective* connection (after any server/db promotion).
            effective_conn = (
                self._credential_store.apply_overrides(conn)
                if self._credential_store else conn
            )

            entry: Dict[str, Any] = {
                "datasource": ds.get("caption") or ds.get("name", ""),
                "class":      cls,
                "server":     effective_conn.get("server", ""),
                "database":   effective_conn.get("dbname", ""),
            }
            if effective_conn.get("port"):
                entry["port"] = effective_conn["port"]
            if effective_conn.get("schema"):
                entry["schema"] = effective_conn["schema"]
            if effective_conn.get("warehouse"):
                entry["warehouse"] = effective_conn["warehouse"]
            # Databricks-specific connection metadata
            if effective_conn.get("http_path"):
                entry["http_path"] = effective_conn["http_path"]
            if effective_conn.get("catalog"):
                entry["catalog"] = effective_conn["catalog"]
            if effective_conn.get("connector_function"):
                entry["connector_function"] = effective_conn["connector_function"]

            # Credential information (from the credentials file).
            # Credential information (from the credentials file).
            if self._credential_store:
                cred = self._credential_store.match(conn)
                if cred:
                    if cred.username:
                        entry["username"] = cred.username

                    entry["has_password"] = bool(cred.password)

                    # Databricks PAT is not written to PBIP/TMDL.
                    # This flag tells users/deployment automation that
                    # PAT authentication needs to be configured securely.
                    if getattr(cred, "token", ""):
                        entry["authentication"] = "PersonalAccessToken"
                        entry["has_personal_access_token"] = True

                    entry["credentials_matched"] = True
                else:
                    entry["credentials_matched"] = False
                    entry["warning"] = (
                        f"No credentials entry matched class='{cls}' "
                        f"server='{conn.get('server', '')}'"
                    )

            manifest["datasources"].append(entry)

        if not manifest["datasources"]:
            return  # No live connections — nothing to write.

        out_path = self.output / "credentials_manifest.json"
        with open(out_path, "w", encoding="utf-8") as fh:
            json.dump(manifest, fh, indent=2)
        matched = sum(
            1 for e in manifest["datasources"] if e.get("credentials_matched")
        )
        total = len(manifest["datasources"])
        _log_creds.info(
            f"credentials_manifest.json -> {out_path}  "
            f"({matched}/{total} datasources matched)"
        )

    def _clean_generated_output(self) -> None:
            """Remove or rename stale generated PBIP output before writing fresh files.

            Windows sometimes fails with:
                OSError: [WinError 145] The directory is not empty

            especially for SemanticModel/data/federated... extract folders.
            If deletion fails, rename the stale folder and continue so the new
            conversion can create a fresh output folder.
            """
            if not self.output.exists():
                return

            # Safety guard: only clean generated PBIP-looking folders.
            if not self.output.name.lower().endswith("_pbip"):
                print(
                    f"  [CLEAN] Skipping cleanup because output folder does not "
                    f"end with '_pbip': {self.output}"
                )
                return

            _log_clean.info(f"Removing stale generated PBIP output: {self.output}")

            def _onerror(func, path, exc_info):
                try:
                    os.chmod(path, stat.S_IWRITE)
                    func(path)
                except Exception:
                    pass

            last_exc = None

            for attempt in range(1, 7):
                try:
                    gc.collect()
                    shutil.rmtree(self.output, onerror=_onerror)
                    return
                except OSError as exc:
                    last_exc = exc
                    time.sleep(0.5 * attempt)

            # If delete still fails, rename stale folder and continue.
            stale_output = self.output.with_name(
                f"{self.output.name}.__stale_{int(time.time())}"
            )

            try:
                self.output.rename(stale_output)
                print(
                    f"  [CLEAN] Could not delete output folder immediately; "
                    f"renamed it to: {stale_output}"
                )

                # Best-effort cleanup after rename.
                try:
                    shutil.rmtree(stale_output, onerror=_onerror)
                except OSError:
                    print(
                        f"  [CLEAN] Renamed stale folder is still locked. "
                        f"Conversion will continue. Delete manually later: {stale_output}"
                    )

                return

            except Exception as rename_exc:
                raise OSError(
                    f"Could not delete or rename output folder: {self.output}. "
                    f"Close Power BI Desktop, File Explorer, terminals, or any process "
                    f"using this folder, then rerun."
                ) from (last_exc or rename_exc)
        
    def run(self) -> None:
        
        name = Path(self.twb_path).stem
        mode = "TWB (visuals only, stub model)" if self.stub_only else "TWBX (full model + visuals)"
        print(f"Tableau -> PBIP  |  {name}")
        print(f"  Mode:    {mode}")
        print(f"  Source:  {self.twb_path}")
        print(f"  Output:  {self.output}")

        # Clean stale generated artifacts from previous runs.
        # This is especially important when switching from Tableau extract
        # CSV mode to Databricks live mode.
        self._clean_generated_output()


        # ---- 1. parse ----------------------------------------------------
        print("  [1/4] Parsing Tableau workbook...")
        parser = TWBParser(self.twb_path)
        parser.parse()
        print(f"        datasources={len(parser.datasources)}  "
              f"parameters={len(parser.parameters)}  "
              f"worksheets={len(parser.worksheets)}  "
              f"dashboards={len(parser.dashboards)}")

        # Optionally dump IR JSON. This is a debug aid — useful when a
        # visual ends up looking wrong and the user wants to see exactly
        # what the parser pulled out.
        if self.debug_ir:
            self.output.mkdir(parents=True, exist_ok=True)
            write_json(self.output / "_ir.json", _build_ir(parser))

# ---- 2. model ----------------------------------------------------
        print("  [2/4] Building semantic model...")
        hyper_data_by_ds = self._bind_hyper_files(parser.datasources)
        model = SemanticModel(
            parser.datasources,
            parameters=parser.parameters,
            stub_only=self.stub_only,
            hyper_data_by_ds=hyper_data_by_ds,
            worksheets=parser.worksheets,
            credential_store=self._credential_store,
        )
        model.build()
        total_cols = sum(len(t["columns"]) for t in model.tables)
        total_measures = sum(len(t.get("measures", [])) for t in model.tables)
        param_table = any(t["name"] == "Parameters" for t in model.tables)
        print(f"        tables={len(model.tables)}  columns={total_cols}  "
              f"relationships={len(model.relationships)}  "
              f"measures={total_measures}  params={'yes' if param_table else 'no'}")

        # ---- 3. report ---------------------------------------------------
        print("  [3/4] Building report pages and visuals...")
        rb = ReportBuilder(parser.datasources, parser.worksheets,
                          parser.dashboards, model, hints=self.hints)
        pages = rb.build()
        from collections import Counter
        type_dist = Counter(
            v["visual"]["visualType"] for p in pages for v in p.get("visuals", [])
        )
        print(f"        pages={len(pages)}  "
              f"visuals={sum(len(p.get('visuals', [])) for p in pages)}")
        print(f"        types={dict(type_dist)}")

        # ---- 4. write ----------------------------------------------------
        print("  [4/4] Writing PBIP files...")
        PBIPWriter(self.output, name).write(model, pages)

        # Write datasource mapping report (helps debug visual field issues)
        self._write_mapping_report(parser.datasources, model)

        # Write credentials manifest (documents connection credentials for
        # PBI Desktop / REST API setup).
        self._write_credentials_manifest(parser.datasources)

        print(f"  Done. Open {self.output / (name + '.pbip')} in Power BI Desktop.")


# ---------------------------------------------------------------------------
# Intermediate representation (parser dicts -> JSON-friendly snapshot)
# ---------------------------------------------------------------------------

def _build_ir(parser: TWBParser) -> Dict[str, Any]:
    """Build a JSON-serializable view of the parsed workbook.

    Mostly a 1:1 copy of the parser's dicts — written to disk so the user
    can inspect what was extracted and see why a visual landed where it
    did. This is the "XML -> JSON" stage the user asked about; everything
    downstream of here is just JSON-to-PBIP transformation.
    """
    return {
        "source_file": parser.twb_path,
        "datasources": [
            {
                "name":          ds["name"],
                "caption":       ds["caption"],
                "connection":    ds["connection"],
                "extracts":      ds.get("extracts", []),
                "extractFilters": ds.get("extractFilters", []),
                "dataSourceFilters": ds.get("dataSourceFilters", []),
                # Stringify (table, col) tuples so json.dump doesn't fail.
                "colsMap":       {k: list(v) for k, v in
                                  (ds.get("colsMap") or {}).items()},
                "objects":       ds["objects"],
                "columns":       ds["columns"],
                "relationships": ds["relationships"],
            }
            for ds in parser.datasources
        ],
        "parameters": parser.parameters,
        "worksheets": parser.worksheets,
        "dashboards": parser.dashboards,
    }


# ---------------------------------------------------------------------------
# .twbx support
# ---------------------------------------------------------------------------

def extract_twbx(
    twbx_path: str, dest_dir: str,
) -> Tuple[str, List[str]]:
    """Pull the .twb and all .hyper files from a .twbx archive.

    Returns (twb_path, [hyper_path, ...]).
    """
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)

    twb_path:    Optional[str] = None
    hyper_paths: List[str]     = []

    with zipfile.ZipFile(twbx_path, "r") as zf:
        for member in zf.namelist():
            if member.endswith(".twb"):
                zf.extract(member, dest_dir)
                twb_path = str(dest / member)
            elif member.endswith(".hyper"):
                zf.extract(member, dest_dir)
                hyper_paths.append(str(dest / member))

    if not twb_path:
        raise ValueError(f"No .twb found inside {twbx_path}")

    if hyper_paths:
        _log_hyper.info(f"Found {len(hyper_paths)} hyper file(s): "
                        + ", ".join(Path(p).name for p in hyper_paths))

    return twb_path, hyper_paths


# ---------------------------------------------------------------------------
# CLI / public entry point
# ---------------------------------------------------------------------------

def run(
    input_path:       Optional[str] = None,
    output:           Optional[str] = None,
    credentials_path: Optional[str] = None,
    *,
    twb:  Optional[str] = None,   # legacy
    twbx: Optional[str] = None,   # legacy
) -> None:
    """Convert a Tableau workbook to PBIP.

    The mode is auto-detected from the file extension:
        .twb   -> visuals-only (stub data model)
        .twbx  -> full data model + visuals

    Either pass `input_path=` directly, or use the legacy `twb=` /
    `twbx=` keyword arguments (kept for the existing CLI shim).

    Pass `credentials_path=` to supply a JSON or XLSX credentials file
    that overrides connection parameters (server / database / port /
    schema) and populates a ``credentials_manifest.json`` alongside the
    PBIP output.
    """
    src = input_path or twbx or twb
    if not src:
        raise ValueError("No input file given.")
    if not os.path.exists(src):
        raise FileNotFoundError(f"Input not found: {src}")

    ext = Path(src).suffix.lower()
    if ext == ".twbx":
        stem             = Path(src).stem
        work_dir         = str(Path(src).parent / f"_{stem}_extracted")
        twb_path, hypers = extract_twbx(src, work_dir)
        out_path         = output or str(Path(src).parent / f"{stem}_pbip")
        Converter(
            twb_path, out_path, stub_only=False, hyper_paths=hypers,
            credentials_path=credentials_path,
        ).run()
    elif ext == ".twb":
        Converter(
            src, output, stub_only=True,
            credentials_path=credentials_path,
        ).run()
    else:
        raise ValueError(f"Unrecognized extension {ext!r} (expected .twb or .twbx).")
