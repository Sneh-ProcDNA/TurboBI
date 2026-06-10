"""Credentials loader for Qlik -> PBIP conversion.

Reads a JSON or XLSX credentials file and matches entries to Qlik
tables so the converter can emit live database partitions
(`Sql.Database`, `PostgreSQL.Database`, `Snowflake.Databases`,
`DatabricksMultiCloud.Catalogs`) instead of empty stubs / CSV imports.

Two outputs from a matched credential:

  1. Connection overrides -- the server / database / port / schema
     fields are attached to the corresponding model table as a
     ``connection`` dict, which ``partition_m.render_partition_m``
     consumes to emit the live M expression.
  2. Credentials manifest -- ``converter.py`` writes a
     ``credentials_manifest.json`` alongside the PBIP output documenting
     the username (and whether a password was supplied) for each live
     data source so the user knows what to enter in
     Power BI Desktop > Transform Data > Data source settings, or push
     via the Power BI REST API.

Passwords / PATs are intentionally NOT embedded in the emitted M.
PBIP's TMDL format has no secure in-file credential store; embedding
plaintext secrets there exposes them in any version-controlled repo.
The credentials manifest is the correct hand-off point for an
automated deployment script.

--------------------------------------------------------------------
JSON format
--------------------------------------------------------------------

    {
      "connections": [
        {
          "class":    "postgres",
          "server":   "prod-pg.mycompany.com",
          "database": "analytics",
          "port":     "5432",
          "schema":   "public",
          "username": "pbi_reader",
          "password": "s3cr3t"
        },
        {
          "class":     "databricks",
          "server":    "adb-1234.cloud.databricks.com",
          "http_path": "/sql/1.0/warehouses/abc",
          "catalog":   "analytics",
          "schema":    "default",
          "token":     "dapi..."
        }
      ]
    }

  ``class`` is the connector identifier (sqlserver / postgres /
  snowflake / databricks / redshift). ``server`` is optional -- when
  omitted the entry matches any server of that class.

  Optional ``table`` field pins a credential entry to a specific Qlik
  table name (case-insensitive); when present, it wins over class-only
  matches for that table.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from ._logging import get_logger

_log_creds = get_logger("CREDS")


class CredentialEntry:
    """One row of the credentials file."""

    _IDENTITY_KEYS = frozenset({
        "class", "server", "host", "database", "dbname", "db",
        "port", "schema", "warehouse",
        "table", "tables", "qlik_table",
        "http_path", "httppath", "http path",
        "catalog",
        "connector_function", "connector",
        "query", "custom_sql", "sql",
        "default",
    })

    _AUTH_KEYS = frozenset({
        "username", "user",
        "password", "pwd",
        "token", "pat", "access_token", "personal_access_token",
    })

    def __init__(self, data: Dict[str, Any]) -> None:
        def _s(key: str, *aliases: str) -> str:
            for k in (key, *aliases):
                v = data.get(k)
                if v is not None and str(v).strip():
                    return str(v).strip()
            return ""

        self.cls       = _s("class").lower()
        self.server    = _s("server", "host")
        self.database  = _s("database", "dbname", "db")
        self.port      = _s("port")
        self.schema    = _s("schema")
        self.warehouse = _s("warehouse")
        self.username  = _s("username", "user")
        self.password  = _s("password", "pwd")

        # Databricks-specific fields
        self.http_path = _s("http_path", "httppath", "http path")
        self.catalog   = _s("catalog")
        self.token     = _s("token", "pat", "access_token", "personal_access_token")
        self.connector_function = _s("connector_function", "connector")

        # Optional table-pinning. Lets the operator say "table X uses
        # credentials entry Y" when class-only matching is ambiguous
        # (e.g. an app pulls from postgres for one table and snowflake
        # for another and the loadmodel doesn't carry that detail).
        tables_raw = _s("table", "tables", "qlik_table")
        self.tables: List[str] = (
            [t.strip().lower() for t in tables_raw.split(",") if t.strip()]
            if tables_raw else []
        )

        # ``default: true`` marks an entry as the catch-all for any
        # table not pinned via ``tables``. Useful when most tables come
        # from one source (e.g. postgres) and only a few are pinned to
        # another (e.g. databricks).
        default_raw = _s("default")
        self.is_default: bool = default_raw.lower() in {"1", "true", "yes", "y"}

        # Custom SQL override (rare in Qlik but supported for parity
        # with the TurboBI Tableau converter).
        self.query = _s("query", "custom_sql", "sql")

        self.extra: Dict[str, str] = {
            k: str(v).strip()
            for k, v in data.items()
            if k.lower() not in (self._IDENTITY_KEYS | self._AUTH_KEYS)
            and v is not None
            and str(v).strip()
        }

    # ------------------------------------------------------------------
    def apply_to(self, conn: Dict[str, Any]) -> Dict[str, Any]:
        """Return a copy of conn with this entry's overrides applied.

        Security note: passwords / PATs are NOT copied -- the M
        expression must not embed them. They live only in the
        credentials manifest for the user to wire up in PBI Desktop.
        """
        result = dict(conn)

        if self.cls:
            result["class"] = self.cls
        if self.server:
            result["server"] = self.server
        if self.database:
            result["dbname"] = self.database
        if self.port:
            result["port"] = self.port
        if self.schema:
            result["schema"] = self.schema
        if self.warehouse:
            result["warehouse"] = self.warehouse

        if self.http_path:
            result["http_path"] = self.http_path
        if self.catalog:
            result["catalog"] = self.catalog
        if self.connector_function:
            result["connector_function"] = self.connector_function

        if self.query:
            result["customSql"] = [{"name": "credentials override",
                                    "sql": self.query}]

        for k, v in self.extra.items():
            result[k] = v

        return result


class CredentialStore:
    """Ordered list of :class:`CredentialEntry` objects."""

    def __init__(self, entries: List[CredentialEntry]) -> None:
        self._entries = entries

    # ------------------------------------------------------------------
    def match(
        self,
        table_name: str = "",
        cls: str = "",
        server: str = "",
        default_class: str = "",
    ) -> Optional[CredentialEntry]:
        """Return the best-matching entry.

        Priority:
          1. Explicit ``table`` pinning in the credentials entry.
          2. class + server exact match (when ``cls`` is given).
          3. class-only match (when ``cls`` or ``default_class`` is given).
          4. Entry marked ``"default": true`` in the credentials file.
          5. Single-entry fallback when only one credential exists.
        """
        tname_lc  = (table_name or "").strip().lower()
        cls_lc    = (cls or "").strip().lower()
        server_lc = (server or "").strip().lower()
        def_cls_lc = (default_class or "").strip().lower()

        # 1. Table pinning
        if tname_lc:
            for e in self._entries:
                if tname_lc in e.tables:
                    return e

        # 2. class + server
        if cls_lc:
            for e in self._entries:
                if e.cls == cls_lc and e.server.lower() == server_lc:
                    return e
            for e in self._entries:
                if e.cls == cls_lc and not e.server:
                    return e

        # 3. ``default_class`` hint (from --connection-class CLI flag).
        # This is the user's way of saying "for any unpinned table, use
        # the postgres entry" when the credentials file contains entries
        # for multiple classes but the Qlik load script doesn't carry
        # connection metadata we can match against.
        if def_cls_lc:
            for e in self._entries:
                if e.cls == def_cls_lc:
                    return e

        # 4. Entry explicitly flagged ``"default": true``.
        for e in self._entries:
            if e.is_default:
                return e

        # 5. Single-entry convenience -- one credentials.json with one
        # entry that should apply to every table.
        if len(self._entries) == 1:
            return self._entries[0]

        return None

    # ------------------------------------------------------------------
    def apply_overrides(
        self,
        conn: Dict[str, Any],
        table_name: str = "",
    ) -> Dict[str, Any]:
        """Return ``conn`` with the matched entry's overrides applied."""
        entry = self.match(
            table_name=table_name,
            cls=(conn.get("class") or "").strip(),
            server=(conn.get("server") or "").strip(),
        )
        return entry.apply_to(conn) if entry is not None else conn

    @property
    def entries(self) -> List[CredentialEntry]:
        return list(self._entries)

    def is_empty(self) -> bool:
        return not self._entries


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_credentials(path: str | Path) -> CredentialStore:
    """Load a credentials file. Supports ``.json`` and ``.xlsx`` / ``.xls``."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Credentials file not found: {path}")

    suffix = p.suffix.lower()
    if suffix == ".json":
        return _load_json(p)
    if suffix in (".xlsx", ".xls"):
        return _load_xlsx(p)
    raise ValueError(
        f"Unsupported credentials file format: {suffix!r} (use .json or .xlsx)"
    )


def _load_json(path: Path) -> CredentialStore:
    with open(path, encoding="utf-8") as fh:
        raw = json.load(fh)

    if isinstance(raw, list):
        rows = raw
    else:
        rows = (
            raw.get("connections")
            or raw.get("credentials")
            or raw.get("datasources")
            or []
        )

    entries = [CredentialEntry(r) for r in rows if isinstance(r, dict)]
    _log_creds.info(f"Loaded {len(entries)} credential entries from {path.name}")
    return CredentialStore(entries)


def _load_xlsx(path: Path) -> CredentialStore:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read XLSX credentials. "
            "Install with:  pip install openpyxl"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Credentials"] if "Credentials" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return CredentialStore([])

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    entries: List[CredentialEntry] = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        entry_data: Dict[str, Any] = {}
        for h, v in zip(headers, row):
            if h and v is not None and str(v).strip():
                entry_data[h] = str(v).strip()
        if entry_data:
            entries.append(CredentialEntry(entry_data))

    _log_creds.info(f"Loaded {len(entries)} credential entries from {path.name}")
    return CredentialStore(entries)


# ---------------------------------------------------------------------------
# Manifest writer
# ---------------------------------------------------------------------------

def write_credentials_manifest(
    out_path: Path,
    bindings: List[Dict[str, Any]],
) -> None:
    """Write ``credentials_manifest.json`` next to the PBIP output.

    Each entry in ``bindings`` is a dict produced by the converter:

        {
          "table":    "<PBI table name>",
          "class":    "postgres",
          "server":   "...",
          "database": "...",
          "schema":   "...",
          "username": "...",
          "password_supplied": bool,
          "token_supplied":    bool,
        }

    Passwords and tokens are NOT serialised -- only a boolean flag so
    the user knows whether they need to enter a secret in PBI Desktop.
    """
    payload = {
        "note": (
            "These data sources need credentials wired up in Power BI "
            "Desktop > Transform Data > Data source settings, or pushed "
            "via the Power BI REST API. Passwords / tokens are NOT "
            "embedded in the PBIP for security."
        ),
        "datasources": bindings,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    _log_creds.info(f"Wrote credentials manifest: {out_path}")
